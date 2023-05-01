# Import necessary libraries
from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
import os
import time
import ctypes

def start():
    # Set the URL and list of indexes to exclude
    url = 'https://enr-apps.as.cmu.edu/open/SOC/SOCServlet/search'
    minus = [4, 5, 9, 10, 11]

    # Set up the webdriver options and initialize a browser object
    options = webdriver.ChromeOptions()
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    browser = webdriver.Chrome(options=options)
    browser.get(url)

    # Select the program location and submit the form
    select = Select(browser.find_element(By.NAME, 'PRG_LOCATION'))
    select.select_by_value("DOH")
    browser.find_element(By.NAME, "SUBMIT").click()
    html = browser.page_source

    # Parse the HTML using BeautifulSoup and extract the titles and tables
    soup = BeautifulSoup(html, 'html.parser')
    titles = []
    for title in soup.find_all('h4'):
        titles.append(title.get_text(strip=True))
    tables = []
    for table in soup.find_all('table'):
        tables.append(table)

    cnt = 0
    dir = os.path.join(os.path.dirname(__file__), 'Databases')
    if not os.path.exists(dir):
        os.makedirs(dir)

    # Iterate over each table, convert it to a dataframe and export it to an Excel file
    for table in tables:
        cnt = cnt + 1
        df = pd.read_html(str(table))[0]

        # Create a new Excel workbook and worksheet
        workbook = Workbook()
        worksheet = workbook.active
        
        # Write the column titles to the worksheet
        col = 1
        for col_num, column_title in enumerate(df.columns, 1):
            if col_num in minus:
                continue
            cell = worksheet.cell(row=1, column=col)
            col = col + 1
            cell.value = column_title

        # Write the data rows to the worksheet
        for row_num, row_data in enumerate(df.values, 2):
            col = 1
            for col_num, cell_data in enumerate(row_data, 1):
                if col_num in minus:
                    continue
                cell = worksheet.cell(row=row_num, column=col)
                col = col + 1
                cell.value = cell_data
        
        # Save the workbook to an Excel file with the same name as the title in the "Databases" folder
        filename = titles[cnt] + '.xlsx'
        path = os.path.join(dir, filename)
        print('Finished table ' + titles[cnt])
        # ctypes.windll.user32.MessageBoxW(0, "Exported into table " + titles[cnt], "Finished job", 1)
        workbook.save(path)

    # Close the browser and print completion message
    time.sleep(1)
    browser.quit()
    print("Done!")
